'''WeDo Modulu / WeeZard'''

'''@Author
Name: Esin Ece Aydin / Joshgun Rzabayli
Date: 24.06.2019 - 18.07.2019
Intrapreneurship Mazars Denge'''

'''brut - net maas hesaplama'''

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

DirecPath = "Salary/scripts/"



class Employee:
    "This is the class for each employee"
    # type of salaries
    __gross = 0.0
    __net = 0.0
    # these are flags to determine convert type
    __gross_to_net = 0
    __net_to_gross = 0
    # minimum and maximum values of basic fee. a salary with smaller than min or larger than max is not allowed
    __basic_fee_min = 0.0
    __basic_fee_max = 0.0
    # features about income tax
    __income_tax_pct = 0.0
    __income_tax = 0.0

    __prim = 0.0
    __stamp_tax = 0.0
    # agi is added to net salary, then we'll find total salary
    __agi_min = 0.0
    __agi_max = 0.0
    # needed variables to define agi
    __kids_count = 0
    __marital_status = 0
    __partner_status = 0
    __boss_cost = 0
    __sale_5 = 0
    # after determining agi, we obtain total salary
    __total_salary = 0.0
    # dictionaries those will be used calculating income tax
    __limit = {}
    __pct = {}
    __agi = 0.0
    data = {}
    # dictionaries for store values month by month
    __income_tax_base = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}
    __cumul_base = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}
    __salary = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}

    # Functions of 'Employee' class ------------------------------------------------------------
    # parameters that are written in the file is read
    def readParameter(self, line_number):
        file = open(DirecPath+'parametre2019.txt', mode='r')  # a file with name parametre2019.txt is opened
        lines = file.read().split('\n')  # split lines
        words = lines[line_number].split(':')  # split words which divided by ':'
        value = float(words[1])  # assign the value that comes after ':'
        return value  # the value in the line_number is returned

    # initializer / instance attributes
    def __init__(self, kids_count, marital_status, partner_status, boss_cost, sale_5, gross_to_net, net_to_gross,salaries_dict):
        self.__gross_to_net = gross_to_net
        self.__net_to_gross = net_to_gross
        self.__boss_cost = boss_cost
        self.__sale_5 = sale_5
        self.__marital_status = marital_status
        self.__partner_status = partner_status
        self.__kids_count = kids_count
        self.defineAgi()
        self.readSalaries(salaries_dict)

        # these are the parameters read from parameter.txt file:
        self.stamp_tax_pct = self.readParameter(1)
        self.sgk_boss_emp_pct = self.readParameter(2)
        self.sgk_boss_unemp_pct = self.readParameter(3)
        self.sgk_emp_pct = self.readParameter(4)
        self.sgk_unemp_pct = self.readParameter(5)
        self.__basic_fee_min = self.readParameter(6)
        self.__basic_fee_max = self.readParameter(7)


    def readSalaries(self,salaries_dictionary):

        temp = 0.0
        for i in range(1,13): # for each month
            if i in salaries_dictionary: # if the month's salary is given
                temp = salaries_dictionary[i]
                continue
            else:
                if temp == 0.0: # if previous month salaries are not given
                    salaries_dictionary[i] = 0.0
                else:
                    salaries_dictionary[i] = temp # set it to the value of the previous month

        self.__salary.update(salaries_dictionary)


    # to get salary from employee. It's invoked in saveToFile func for 12 times

    # a function to calculate total prim
    def totalPrim(self):
        # if gross is larger than max basic fee, it won't calculate prim according to gross amount.
        if self.__gross > self.__basic_fee_max:
            emp_prim = self.__basic_fee_max * self.sgk_emp_pct
            unemp_prim = self.__basic_fee_max * self.sgk_unemp_pct
        else:
            emp_prim = self.__gross * self.sgk_emp_pct
            unemp_prim = self.__gross * self.sgk_unemp_pct

        self.__prim = emp_prim + unemp_prim

    # read percentages and limits from the file with name as gelirvergisi2019.txt
    def readIncomeTaxPct(self):
        file = open(DirecPath+'gelirvergisi2019.txt', mode='r')
        lines = file.read().split('\n')
        for line_number in range(4):  # there are always 4 limit. And txt file also has 4 lines.
            words = lines[line_number].split(':')  # split the words in an exact line_number divided by ':'
            self.__limit[line_number] = int(words[0])  # 0 means limit value
            self.__pct[line_number] = float(words[1])  # 1 means percantage value

        file.close()

    # a function to calculate total income tax
    # when income pie is changing the tax percentage is defined for 2 pies separately
    def totalIncomeTax(self, month):
        self.calculateCumulBase(month)  # a function is invoked to calculate cumulative basea
        self.readIncomeTaxPct()  # a func is invoked to read percentages from file
        # limits read from file are assigned to local variables to shorten.
        limit1 = self.__limit[0]  # 0
        limit2 = self.__limit[1]  # 18000
        limit3 = self.__limit[2]  # 40000
        limit4 = self.__limit[3]  # 98000
        __income_tax = 0.0  # firstly local income_tax is assigned to 0
        if self.__cumul_base[month] > limit1 and self.__cumul_base[month] < limit2:  # 0-18
            # print("vergi dilimi = ", self.pct[0] )
            self.__income_tax_pct = self.__pct[0]
            __income_tax = self.__income_tax_base[month] * self.__pct[0]

        elif self.__cumul_base[month] > limit2 and self.__cumul_base[month] < limit3:  # 18-40
            # print("vergi dilimi = ", self.pct[1])

            if self.__cumul_base[month - 1] < limit2:
                __income_tax += (limit2 - self.__cumul_base[month - 1]) * self.__pct[0]
                __income_tax += (self.__cumul_base[month] - limit2) * self.__pct[1]
                print("tax pie is changing 18-40")
            else:
                __income_tax = self.__income_tax_base[month] * self.__pct[1]
                self.__income_tax_pct = self.__pct[1]

        elif self.__cumul_base[month] > limit3 and self.__cumul_base[month] < limit4:  # 40-148
            # print("vergi dilimi = ", self.pct[2] )
            self.__income_tax_pct = self.__pct[2]
            if self.__cumul_base[month - 1] < limit3:
                __income_tax += (limit3 - self.__cumul_base[month - 1]) * self.__pct[1]
                __income_tax += (self.__cumul_base[month] - limit3) * self.__pct[2]
                print("tax pie is changing 40-148")
            else:
                __income_tax = self.__income_tax_base[month] * self.__pct[2]
                self.__income_tax_pct = self.__pct[2]

        elif self.__cumul_base[month] > limit4:  # 148- ...
            # print("vergi dilimi = ", self.pct[3] )
            self.__income_tax_pct = self.__pct[3]
            if self.__cumul_base[month - 1] < limit4:
                __income_tax += (limit4 - self.__cumul_base[month - 1]) * self.__pct[2]
                __income_tax += (self.__cumul_base[month] - limit4) * self.__pct[3]
                print("tax pie is changing 148-")
            else:
                __income_tax = self.__income_tax_base[month] * self.__pct[3]
                self.__income_tax_pct = self.__pct[3]
        return __income_tax

    def calculateCumulBase(self, month):
        base = {}  # an empty dict to store income base temporarily
        base[month] = self.__gross - self.__prim
        self.__cumul_base.update(base)  # cumulative base dict is updated according to base dict
        self.__income_tax_base.update(base)  # it's not cumulative, it stores only the income base for each month
        if month == 1:  # if month is January
            self.__cumul_base[month] = self.__cumul_base[month]
        else:  # if month is other than January
            self.__cumul_base[month] += self.__cumul_base[month - 1]
        # we'll need zero index later.
        zero_index = {}
        zero_index[0] = 0.0
        self.__cumul_base.update(zero_index)

    # to calculate total stamp tax
    def totalStampTax(self, month):
        self.__stamp_tax = self.__gross * self.stamp_tax_pct

    # read value from the file with name agi2019.txt
    def readAgi(self, line_number):
        file = open(DirecPath+'agi2019.txt', mode='r')
        lines = file.read().split('\n')  ##split lines
        words = lines[line_number].split(':')  # split words which divided by ':'
        value = float(words[1])  # the name is written at the 0 index. the value is written at the 1 index
        return value

    # a function to determine which agi should be selected for information of employee
    def defineAgi(self):
        self.__agi_min = self.readAgi(0)
        self.__agi_max = self.readAgi(1)


        if (self.__marital_status == 0):
            self.__agi = self.readAgi(2)
        elif (self.__marital_status == 1 and self.__partner_status == 0):
            if self.__kids_count == 0:
                self.__agi = self.readAgi(3)
            elif self.__kids_count == 1:
                self.__agi = self.readAgi(4)
            elif self.__kids_count == 2:
                self.__agi = self.readAgi(5)
            elif self.__kids_count == 3:
                self.__agi = self.readAgi(6)
            elif self.__kids_count == 4:
                self.__agi = self.readAgi(7)
            elif self.__kids_count >= 5:
                self.__agi = self.readAgi(8)
            else:
                print("a problem arise at this moment")
        elif (self.__marital_status == 1 and self.__partner_status == 1):
            if self.__kids_count == 0:
                self.__agi = self.readAgi(9)
            elif self.__kids_count == 1:
                self.__agi = self.readAgi(10)
            elif self.__kids_count == 2:
                self.__agi = self.readAgi(11)
            elif self.__kids_count == 3:
                self.__agi = self.readAgi(12)
            elif self.__kids_count == 4:
                self.__agi = self.readAgi(13)
            elif self.__kids_count >= 5:
                self.__agi = self.readAgi(14)
            else:
                print("a problem arise at this moment")
        # return agi

    # a function to invoke other needed functions to calculate gross or net salary
    def calculate(self, month):
        if self.__gross_to_net == 1:  # if net salary is calculated
            self.__net = self.calculateNet(month)
        elif self.__net_to_gross == 1:  # if gross salary is calculated
            self.__net = self.__salary[month]

            prim_pct = self.sgk_emp_pct + self.sgk_unemp_pct
            self.readIncomeTaxPct() #a func is invoked to read percentages from file
            if month == 1:
                if self.__net < 21176.470588235294:
                    self.__income_tax_pct = self.__pct[0]
                elif self.__net > 21176.470588235294 and self.__net < 47.05882352941177:
                    self.__income_tax_pct = self.__pct[1]
                elif self.__net > 47.05882352941177 and self.__net < 174.11764705882354:
                    self.__income_tax_pct = self.__pct[2]
                elif self.__net > 174.11764705882354:
                    self.__income_tax_pct = self.__pct[3]
                self.__gross = (self.__net) / (1 - prim_pct - self.stamp_tax_pct - self.__income_tax_pct + (prim_pct * self.__income_tax_pct))
            else:
                self.__gross = (self.__net + self.__income_tax ) / (1 - prim_pct - self.stamp_tax_pct)


            self.totalPrim()
            self.__income_tax = self.totalIncomeTax(month)
            self.totalStampTax(month)
            self.__total_salary = self.__gross - self.__prim - self.__income_tax - self.__stamp_tax + self.__agi
            # self.__total_salary = self.__net + self.__prim + self.__income_tax + self.__stamp_tax - self.__agi

    def approxiamateNet(self, approach_net, month):
        iter = 0
        approach_net = self.calculateNet(month)
        while (self.__net > approach_net - 10 or self.__net < approach_net + 10) or iter < 500:
            self.calculateNet(month)
            iter += 1
            if iter > 500:
                print("break oldu")
                break

    def calculateNet(self, month):
        self.__gross = self.__salary[month]
        if self.__net_to_gross == 1:  # if gross salary is calculated
            self.__gross += self.__basic_fee_min
            print(self.__gross)

        self.totalPrim()
        self.__income_tax = self.totalIncomeTax(month)
        self.totalStampTax(month)
        # if income tax is smaller than agi min, agi is determined as much as income tax
        if (self.__income_tax < self.__agi_min and self.__income_tax != 0):
            self.__agi = self.__income_tax
        # if income tax is larger than agi max, agi is determined as much as income tax
        """elif(self.__income_tax > self.__agi_max):
            self.__agi = self.__agi_max
        """
        net = self.__gross - self.__prim - self.__income_tax - self.__stamp_tax
        self.__total_salary = self.__gross - self.__prim - self.__income_tax - self.__stamp_tax + self.__agi
        return net

    # a function to open a file with name hesaplama.txt to write just titles
    def openFile(self, __gross_to_net, __net_to_gross):
        file = open('hesaplama.txt', mode='w')
        if __gross_to_net == 1:
            file.write(
                "Month\tGross\t\tTotal Prim\tCumulative Income Base\tIncome Tax\t\tStamp Tax\t\tNet\t\tAgi\t\tTotal Salary\tTotal Cost for Employer\n")
        elif __net_to_gross == 1:
            file.write(
                "Month\tNet\t\tTotal Prim\tCumulative Income Base\tIncome Tax\t\tStamp Tax\t\tGross\t\tAgi\t\tTotal Salary\tTotal Cost for Employer\n")
        file.close()

    # a function to save results to a file with name hesaplama.txt
    def saveToFile(self):
        self.openFile(self.__gross_to_net, self.__net_to_gross)  # a func is invoked to prepare file
        file = open('hesaplama.txt', mode='a')
        self.month_names = {
            1: "Jan",
            2: "Feb",
            3: "Mar",
            4: "Apr",
            5: "May",
            6: "June",
            7: "July",
            8: "Aug",
            9: "Sep",
            10: "Oct",
            11: "Nov",
            12: "Dec"
        }





        for month in self.__salary.keys():  # it goes for 12 times (12 months)
            self.calculate(month)  # a func is invoked to calculate for each months salary
            self.data[month] = [ self.__prim, self.__cumul_base[
                    month], self.__income_tax, self.__stamp_tax, self.__agi, self.__total_salary,
                self.findBossCost(month)]
                # 0 and 5th index
            if self.__gross_to_net == 1:
                file.write("%s\t%f\t%f\t%f\t\t%f\t\t%f\t\t%f\t%f\t%f\t%f\n" % (
                self.month_names[month], self.__gross, self.__prim, self.__cumul_base[
                    month], self.__income_tax, self.__stamp_tax, self.__net, self.__agi, self.__total_salary,
                self.findBossCost(month)))
                self.data[month].insert(0,self.__gross)
                self.data[month].insert(5,self.__net)
            elif self.__net_to_gross == 1:
                file.write("%s\t%f\t%f\t%f\t\t%f\t\t%f\t\t%f\t%f\t%f\t%f\n" % (
                self.month_names[month], self.__net, self.__prim, self.__cumul_base[
                    month], self.__income_tax, self.__stamp_tax, self.__gross, self.__agi, self.__total_salary,
                self.findBossCost(month)))
                self.data[month].insert(0,self.__net)
                self.data[month].insert(5,self.__gross)
        file.close()

        self.saveToExcel()

    def saveToExcel(self):
        print("Now, the results will be written to Excel file")

        wb = Workbook()
        ws = wb.active
        ws.sheet_format.defaultColWidth = 24  # custom width

        """" Title """

        ws.title = "Salary calculation"  # sheet name
        ws.merge_cells('A1:B1')  # merge the cells for the image to fit
        img = Image(DirecPath+"mazars logo.jpg")  # open the image
        img.height = 83
        img.width = 335
        ws.add_image(img, 'A1')
        ws.row_dimensions[1].height = 65
        """"""

        ws.cell(row = 3, column=1, value='Year')
        ws.cell(row = 3, column=2, value='2019')

        ws.cell(row=4,column=1,value='Social status')

        if self.__marital_status == 0:
            ws.cell(row = 4,column=2,value="Single")
        else:
            ws.cell(row = 4, column=2, value="Married")


        ws.cell(row=5,column=1,value="Partner")
        if self.__partner_status == 0:
            ws.cell(row=5,column=2,value="Does not work")
        else:
            ws.cell(row=5,column=2, value="Works")

        ws.cell(row = 6, column=1, value="Kids count")
        ws.cell(row=6, column=2,value=self.__kids_count)

        ws.merge_cells('A7:B7')
        if self.__boss_cost == 1:
            _= ws.cell(row = 7, column=1, value="Total cost for employer is included")

        else:
            _ = ws.cell(row=7,column=1, value="Total cost for employer is not included ")


        ws.merge_cells('A8:C8')
        if self.__sale_5 == 1:
            _ = ws.cell(row = 8, column=1, value="employer's share of social security premium, 5 percent sale is taken into account")
        else:
            _ = ws.cell(row = 8,column=1, value="employer's share of social security premium, 5 percent sale is not taken into account")

        """Calculation  header"""

        ws.merge_cells('A9:L9')
        ws.cell(9, 1).font = Font(bold=True, size=15)
        title_cell = ws.cell(9, 1)
        title_cell.alignment = Alignment(horizontal='center')

        """"""

        """Headers"""


        if self.__gross_to_net == 1:
            headers = ['Month', 'Gross', 'Total Prim', 'Cumulative Income Base', 'Income Tax', 'Stamp Tax',
                   'Net', 'Agi', 'Total Salary', 'Total Cost for Employer']
            ws['A9'] = "Gross to Net"
        elif self.__net_to_gross==1:
            ws['A9'] = "Net to Gross"
            headers = ['Month','Net','Total Prim','Cumulative Income Base','Income Tax','Stamp Tax','Gross','Agi','Total Salary','Total Cost for Employer']


        col_Index = 1  # Start Index

        ws.row_dimensions[10].height = float(33.75)  # row height for the header
        for header in headers:
            _ = ws.cell(row=10, column=col_Index, value=header)
            _.font = Font(bold=True, size=12)
            _.alignment = Alignment(horizontal='center')
            col_Index += 1

        """"""

        data_row = 11


        for month in self.__salary.keys():  # it goes for 12 times (12 months)
            self.calculate(month)  # a func is invoked to calculate for each months salary
            data = [self.__prim, self.__cumul_base
                , self.__income_tax, self.__stamp_tax, self.__net, self.__agi, self.__total_salary]
            data_column = 3
            ws.cell(row=data_row,column=1,value=self.month_names[month])

            for d in data:

                if data_column ==7 : data_column += 1 ; continue

                elif data_column == 4:
                    ws.cell(row=data_row,column=data_column,value=self.__cumul_base[month])
                    data_column += 1
                    continue
                ws.cell(row=data_row,column=data_column,value=d)
                data_column += 1

            ws.cell(row=data_row,column=10,value=self.findBossCost(month))

            if self.__gross_to_net == 1:
                ws.cell(row= data_row,column=2, value=self.__gross)
                ws.cell(row= data_row, column=7,value=self.__net)
            elif self.__net_to_gross==1:
                ws.cell(row=data_row,column=2,value=self.__net)
                ws.cell(row = data_row,column=7, value=self.__gross)

            data_row += 1

        wb.save(DirecPath+"ExcelFolders/Mazars_Gross_Net_Calculation.xlsx")


    # a func to find boss cost. if boss_cost flag is 0, it returns 0.0
    def findBossCost(self, month):
        if self.__sale_5 == 1 and month == 1:  # if 5% sale is applied
            self.sgk_boss_emp_pct -= 0.05
        if self.__boss_cost == 1:
            if self.__gross > self.__basic_fee_max:
                boss_share = self.sgk_boss_emp_pct * self.__basic_fee_max

                unemp_prim = self.sgk_boss_unemp_pct * self.__basic_fee_max
            else:
                boss_share = self.sgk_boss_emp_pct * self.__gross
                unemp_prim = self.sgk_boss_unemp_pct * self.__gross
            """
            boss_share = self.sgk_boss_emp_pct * self.__gross
            unemp_prim = self.sgk_boss_unemp_pct * self.__gross"""

            sum = self.__gross + boss_share + unemp_prim
        else:
            sum = 0.0

        return sum


"""
#if gross is larger than max basic fee, it won't calculate prim according to gross amount.
        if self.__gross > self.__basic_fee_max:
            emp_prim = self.__basic_fee_max * self.sgk_emp_pct
            unemp_prim = self.__basic_fee_max * self.sgk_unemp_pct
        else:
            emp_prim = self.__gross * self.sgk_emp_pct
            unemp_prim = self.__gross * self.sgk_unemp_pct

        self.__prim = emp_prim + unemp_prim
"""