from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from pandas import DataFrame, notnull
from openpyxl.styles import Font
from numpy import NaN as NaN
import os
import excel
import datetime

__author__ = "Fahed Sabellioglu"


class Excel:
    __nrows = 0
    __Bakiye = {}
    __Zero_Codes = [191, 192, 391]
    __Ignored_Codes = [590, 591]

    def __init__(self, files, path, saveAs, b_date="", b_name=""):
        self.saveAs = saveAs
        self.files = files
        self.path = path
        self.b_date = b_date
        self.b_name = b_name

    def workbook_to_df(self, file_name):
        """
        :param file_name: the file name that needs to be opened
        :return: returns the excel workbook and the dataframe including the data from the excel workbook
        """
        open_workbook = load_workbook(filename=file_name, read_only=True)  # open the workbook in read only mode
        sheet_name = open_workbook.sheetnames[0]  # get the zero sheet
        ws = open_workbook[sheet_name]  # open the sheet
        excel_data = list(ws.values)  # get the data inside the workbook row by row
        titles = excel_data[0]  # zero index of excel_data is the titles
        df = DataFrame(excel_data[1:], columns=titles)  # create a dataframe with the data excluding titles
        return open_workbook, df

    def bilanco_template(self):
        """
        :return:  the workbook of the template and the both the income and balance sheets
        """
        workbook = load_workbook(self.resource_path('TEMPLATE BS-IS.xlsx'))  # open the template
        Balance_sheet, Income_sheet = workbook.sheetnames[0:2]  # get the first two sheets
        ws_balance = workbook[Balance_sheet]
        ws_income = workbook[Income_sheet]
        return workbook, ws_income, ws_balance

    def bilanco(self):
        # self.__Bakiye.clear()
        direc_path = os.path.dirname(self.path)

        fileName = direc_path + "/" + self.saveAs
        open_workbook, df = self.workbook_to_df(file_name=self.path)
        self.bilanco_calcualtions(df, fileName)

    def loadRules(self):
        """
            opens the rules excel workbook and turns the rules into a dataframe to be checked later on.
        :return: None
        """
        _, self.__rules_df = self.workbook_to_df(self.resource_path('Rules.xlsx'))
        self.deleteColumns([None], self.__rules_df)  # drop the unused rows.
        self.__rules_df.set_index('input', inplace=True)  # set the input codes as index
        self.__rules_df.dropna(inplace=True, how='all')  # drop the rows whose values are all None.

    def bilanco_calcualtions(self, df, saveName):
        """
            a function that calculates the bilanco values and runs the functions to write the values
        :param df: the dataframe that includes the input values.
        :return:
        """
        self.loadRules()  # get the rules of input, output and their output type
        try:
            self.replace_column(['Borç Bakiye', 'Alacak Bakiye'], NaN, 0, df)  # remove nan from the columns
            df['Bakiye'] = df[['Borç Bakiye', 'Alacak Bakiye']].apply(lambda x: x[0] - x[1],
                                                                      axis=1)  # calculate the Bakiye column
            self.deleteColumns(['Detay', 'Borç Toplamı', 'Alacak Toplamı', 'Borç Bakiye', 'Alacak Bakiye'],
                               df)  # delete the unneeded columns
            df = df[notnull(df["Hesap Kodu"])]
            self.codes_balances = dict(
                zip(df['Hesap Kodu'].apply(lambda x: int(x)), df['Bakiye']))  # get codes with their values from Bakiye

        except KeyError:
            raise KeyError(os.path.basename(saveName))

        except TypeError:
            raise TypeError("Something wrong with the format of the cells.")

        for code in self.codes_balances.keys():  # for every code in the input file
            if code not in [950, 951]:  # if the code is not 951 nor 951
                output_code = self.__rules_df.loc[str(code)]['output']
                if output_code not in self.__Bakiye.keys():  # if the output code is not set yet
                    if output_code in self.codes_balances and code != output_code:
                        """
                            if the input code and output code are different from each other 740 622 and already have a value
                            for the 622.
                            620 = absolute value(the value for 740 + the value for 620)
                        """
                        self.__Bakiye[output_code] = abs(self.codes_balances[code] + self.codes_balances[output_code])
                    else:
                        """
                            if the output code and input code are identical 102 102
                            102 = absolute value for 102
                        """
                        self.__Bakiye[output_code] = abs(self.codes_balances[code])
                else:
                    """
                        if an input code is setting a vale to an output code that is already processed earlier.

                        529 = value for 529 + value for 524
                    """
                    self.__Bakiye[output_code] += abs(self.codes_balances[code])

        workbook, ws_income, ws_balance = self.bilanco_template()  # get the template workbook with the balance and income sheets

        " The function writeToSheets explained below. "
        self.writeToSheets(ws_income, 5, 69, ['C'], ['H'])
        self.writeToSheets(ws_balance, 8, 139, ['C', 'O'], ['H', 'Q'])

        """Create a sheet for the Bakiye calculations."""
        BakiyeSheet = workbook.create_sheet("Bakiye")
        BakiyeSheet.column_dimensions['A'].width = 15
        BakiyeSheet.column_dimensions['B'].width = 41
        BakiyeSheet.column_dimensions['C'].width = 15

        self.firm_name([ws_balance['B2'], ws_income['B2']])
        self.firm_date([ws_balance['H4'], ws_income["H3"]])

        self.saveToSheet(df, BakiyeSheet)
        workbook.save(saveName)

    def firm_name(self, sheets):
        for sheet in sheets:
            sheet.value = self.b_name
            sheet.font = Font(bold=True, size=12)

    def firm_date(self, sheets):
        for sheet in sheets:
            sheet.value = "CURRENT PERIOD " + self.b_date

    def writeToSheets(self, sheet, s_range, e_range, i_cols, o_cols):
        """

            Writes to the required sheet by providing the start,end range and the input and the output column.

        :param sheet: the sheets that the values will be written to.
        :param s_range:  the start range in the sheet
        :param e_range:  the end range in the sheet
        :param i_cols: the insert column ( the cell where the codes are written)
        :param o_cols: the output column ( the cell where the output for each input code will be written ).
        :param ignore_codes: the codes that needs to be ignored
        :return:  None
        """
        for x in range(s_range, e_range):
            for in_col, out_col in zip(i_cols, o_cols):  # get an element from both i_col and o_col lists at once.
                col_code = sheet[in_col][x].value  # get the codes value
                self.checker(sheet[out_col][x], col_code)

    def checker(self, sheet, col_val):
        """
            A function that runs in the Bilanco calculations to:
                1- omit the None code cells,
                2- check if the input code has a value in the Bakiye dictionary.
                3- check if the input code is one the codes that needs to be zero.
                    1- if Yes:
                        it will change the cell color to Red.

        :param sheet: the sheet that the values will be written to.
        :param col_val: the input code
        :param ignored_codes: #
        :return:
        """
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        if isinstance(col_val, int):  # omitting the non coders cells
            if col_val not in self.__Ignored_Codes:  # if the codes is not included in the ignore list.
                if col_val in self.__Bakiye:  # if there is an output for the input code
                    if col_val in self.__Zero_Codes and self.__Bakiye[col_val] != 0:  # if the code is one
                        sheet.value = "Not Zero " + str(col_val)
                        sheet.fill = redFill

                    else:
                        sheet.value = self.__Bakiye[col_val]

    def loadFiles(self):
        self.dfs = DataFrame()

        counter = 1
        for file in self.files:
            open_workbook, df = self.workbook_to_df(file_name=file)

            # renaming sube to balance
            df.rename(columns={'Şube': "Balance"}, inplace=True)

            # dropping the unneeded columns
            try:
                self.deleteColumns(['Stok Kodu', 'Sıralama', 'Belge Türü', 'Belge Türü Açıklaması',
                                    'Masraf Merkezi', 'COLUMN1', 'Yevmiye No', 'Fiş Türü'], df)
            except KeyError:
                open_workbook.close()
                raise KeyError(os.path.basename(file))

            df.insert(0, 'identifier', counter)

            # replace 'None' or None with -
            self.replace_column(['Hesap Adı'], ["None", None], '-', df)

            # get only the names of the companies
            filtered_names = df[['Hesap Kodu', 'Hesap Adı']][df['Hesap Adı'] != "-"]

            # get the indexes that includes company names
            delete_indexes = list(filtered_names.index)

            # generate a dictionary Hesap Kodu : Hesap Adi
            names = dict(zip(filtered_names['Hesap Kodu'], filtered_names['Hesap Adı']))

            # fill Hesap Adi with the Hesap Kodu based on the Hesap Kodu
            df['Hesap Adı'] = df['Hesap Kodu'].apply(lambda x: names[x])

            # delete the rows that had the company names
            df.drop(labels=delete_indexes, inplace=True)

            # append the temporary dataframe to the main dataframe
            self.dfs = self.dfs.append(df, ignore_index=True)

            counter += 1

        self.__nrows = len(self.dfs.index)

        # change NaNs into 0 to do the calculations
        self.replace_column(['Borç Tutarı', 'Alacak Tutarı'], NaN, 0, self.dfs)

        # performing Balance = Borc Tutari - Alacak Tutari
        self.dfs["Balance"] = self.dfs[['Borç Tutarı', 'Alacak Tutarı']].apply(lambda x: x[0] - x[1], axis=1)

        self.replace_column(['Borç Tutarı', 'Alacak Tutarı'], NaN, 0, self.dfs)  # remove nan from the columns

        # apply the function => if balance >=0 return doviz tutari else: return 0
        self.dfs["PV Döviz"] = self.dfs[["Balance", 'Döviz Tutarı']].apply(self.pv_doviz, axis=1)

        # replacing Nans with zero
        self.replace_column(["PV Döviz"], NaN, 0, self.dfs)

        # fill C1 and C3 with the manipulation of Hesap Kodu
        account_codes = list(self.dfs['Hesap Kodu'])
        self.dfs.insert(1, 'C1', [value[0] for value in account_codes])
        self.dfs.insert(2, 'C3', [value.split(".")[0] for value in account_codes])

        # format the date from long to short date
        self.dfs[['Fiş Tarihi', 'Evrak Tarihi']] = self.dfs[['Fiş Tarihi', 'Evrak Tarihi']].apply(
            lambda x: x.dt.strftime('%d.%m.%Y'))

        self.dfs = self.dfs.reset_index(drop=True)

        final_workbook = Workbook()
        ws = final_workbook['Sheet']
        ws.title = 'Outcome'

        ws.sheet_format.defaultColWidth = 15  # custom width
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['J'].width = 30

        # save the sheet
        self.saveToSheet(self.dfs, ws)

        self.cells_format(['K', "L", 'M', "O", 'P'], 'Comma', ws)

        path = os.path.join(self.path, self.saveAs)

        self.saveFile(final_workbook, path)

        self.CreatePivot(path)

    def CreatePivot(self, file_path):

        excel.Pivot(self.path, self.saveAs)

    def saveToSheet(self, df, toSheet):
        for row in dataframe_to_rows(df, index=False, header=True):
            toSheet.append(row)

        for cell in toSheet['1:1']:
            cell.font = Font(bold=True, size=12)

    def saveFile(self, ws, path):
        try:
            ws.save(path)
            ws.close()

        except PermissionError:
            raise PermissionError("Please close the file " + self.saveAs + " and try again.")

    def pv_doviz(self, x):
        balance = x[0]
        if (balance >= 0):
            return x[1]
        else:
            return -x[1]

    def time_formart(self, sheet):
        for row in range(2, self.__nrows + 1):
            for col in ['F', 'H']:
                dttm = datetime.datetime.strptime(sheet[col + str(row)].value, "%d/%m/%Y")
                sheet[col + str(row)].value = dttm

    def cells_format(self, column_list, formatType, sheet):
        for row in range(2, self.__nrows + 1):
            for col in column_list:
                sheet[col + str(row)].style = formatType

    def replace_column(self, column_list, value_remove, value_place, df):
        for col in column_list:
            df[col].replace(value_remove, value_place, inplace=True)

    def deleteColumns(self, names_list, df):
        for name in names_list:
            del df[name]

    def resource_path(self, relative_path):
        import sys
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)
